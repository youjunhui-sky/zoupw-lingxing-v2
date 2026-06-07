"""import_owners.py — 从 Excel 导入责任人映射到 slow_moving_owner

用法
----
.. code-block:: bash

    # 1) 干跑 (默认), 只打印不写库
    .venv/bin/python scripts/import_owners.py owners.xlsx

    # 2) 真跑 (--yes)
    .venv/bin/python scripts/import_owners.py owners.xlsx --yes

Excel 列 (大小写不敏感, 顺序不敏感)
-----------------------------------------

==========  =======================  ===================================
列名         必填                     说明
==========  =======================  ===================================
scope        是                       sku / asin / category / shop
scope_value  是                       对应值: SKU 串 / ASIN 串 / 类目 / 店铺 sid
owner_name   是                       责任人姓名
owner_email  否                       邮箱
owner_feishu_open_id 否                V2 推送用
notes        否                       备注
==========  =======================  ===================================

行为
----
* 严格校验 scope 取值 (枚举白名单), 行级错误不中断
* UPSERT 到 slow_moving_owner (ON CONFLICT (scope, scope_value) DO UPDATE)
* 不传 ``--yes`` 时 dry-run: 解析 + 校验 + 打印将发生什么, 不写库
* 输出 summary: 新增 X 条, 更新 Y 条, 跳过(错误) Z 条
"""
from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path
from typing import Any

# 把 src 加到 import path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import create_async_engine

from lingxing.config.settings import settings
from lingxing.models.slow_moving import SlowMovingOwner

VALID_SCOPES = ("sku", "asin", "category", "shop")

# 列名映射 (大小写不敏感, 别名也接受)
COLUMN_ALIASES = {
    "scope": "scope",
    "owner_scope": "scope",
    "scope_value": "scope_value",
    "value": "scope_value",
    "owner": "owner_name",
    "owner_name": "owner_name",
    "name": "owner_name",
    "email": "owner_email",
    "owner_email": "owner_email",
    "feishu_id": "owner_feishu_open_id",
    "open_id": "owner_feishu_open_id",
    "owner_feishu_open_id": "owner_feishu_open_id",
    "notes": "notes",
    "note": "notes",
    "remark": "notes",
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="从 Excel 导入 slow_moving_owner 责任人映射",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("excel", type=Path, help=".xlsx 文件路径")
    p.add_argument("--yes", action="store_true", help="实际写入 (默认 dry-run)")
    p.add_argument("--sheet", type=str, default=None, help="指定 sheet 名 (默认第一个)")
    return p.parse_args()


def load_rows(excel_path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    """读 Excel → list[dict], 自动 normalize 列名"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: 需要 openpyxl. 安装: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows_iter)
    except StopIteration:
        return []

    # normalize header
    header: list[str] = []
    for h in raw_header:
        if h is None:
            header.append("")
            continue
        key = str(h).strip().lower()
        header.append(COLUMN_ALIASES.get(key, key))

    out: list[dict[str, Any]] = []
    for r in rows_iter:
        if all(c is None or str(c).strip() == "" for c in r):
            continue  # skip empty
        row = {header[i]: (r[i] if r[i] is not None else None) for i in range(min(len(header), len(r)))}
        out.append(row)
    return out


def validate_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[tuple[int, str]]]:
    """校验 rows, 返回 (valid, errors[(row_idx, msg)])"""
    valid: list[dict[str, Any]] = []
    errors: list[tuple[int, str]] = []
    for idx, row in enumerate(rows, start=1):  # 1-based row number (含 header)
        scope = (row.get("scope") or "").strip().lower() if row.get("scope") else ""
        scope_value = (row.get("scope_value") or "").strip() if row.get("scope_value") else ""
        owner_name = (row.get("owner_name") or "").strip() if row.get("owner_name") else ""

        if not scope:
            errors.append((idx, "缺少 scope"))
            continue
        if scope not in VALID_SCOPES:
            errors.append((idx, f"scope={scope!r} 非法 (必须是 {VALID_SCOPES})"))
            continue
        if not scope_value:
            errors.append((idx, "缺少 scope_value"))
            continue
        if not owner_name:
            errors.append((idx, "缺少 owner_name"))
            continue

        valid.append({
            "scope": scope,
            "scope_value": str(scope_value),
            "owner_name": str(owner_name),
            "owner_email": (str(row.get("owner_email")).strip() if row.get("owner_email") else None) or None,
            "owner_feishu_open_id": (str(row.get("owner_feishu_open_id")).strip() if row.get("owner_feishu_open_id") else None) or None,
            "notes": (str(row.get("notes")).strip() if row.get("notes") else None) or None,
        })
    return valid, errors


async def do_upsert(rows: list[dict[str, Any]]) -> tuple[int, int]:
    """UPSERT 到 slow_moving_owner, 返回 (inserted, updated)"""
    eng = create_async_engine(settings.postgres_async_url)
    inserted = 0
    updated = 0
    try:
        async with eng.begin() as conn:
            for row in rows:
                # 先查 (scope, scope_value) 是否存在
                r = await conn.execute(
                    text("SELECT id FROM slow_moving_owner WHERE scope=:s AND scope_value=:v"),
                    {"s": row["scope"], "v": row["scope_value"]},
                )
                exists = r.scalar() is not None
                stmt = pg_insert(SlowMovingOwner).values(**row)
                stmt = stmt.on_conflict_do_update(
                    index_elements=["scope", "scope_value"],
                    set_={
                        "owner_name": stmt.excluded.owner_name,
                        "owner_email": stmt.excluded.owner_email,
                        "owner_feishu_open_id": stmt.excluded.owner_feishu_open_id,
                        "notes": stmt.excluded.notes,
                        "sync_time": __import__("datetime").datetime.now(__import__("datetime").timezone.utc),
                    },
                )
                await conn.execute(stmt)
                if exists:
                    updated += 1
                else:
                    inserted += 1
    finally:
        await eng.dispose()
    return inserted, updated


def main() -> None:
    args = parse_args()

    if not args.excel.exists():
        print(f"ERROR: 文件不存在: {args.excel}", file=sys.stderr)
        sys.exit(1)

    print(f"读取 Excel: {args.excel}")
    rows = load_rows(args.excel, args.sheet)
    print(f"  解析到 {len(rows)} 行")

    if not rows:
        print("  无数据, 退出")
        return

    valid, errors = validate_rows(rows)
    print(f"  校验通过: {len(valid)} 条")
    if errors:
        print(f"  校验失败: {len(errors)} 条")
        for row_idx, msg in errors:
            print(f"    - 第 {row_idx} 行: {msg}")

    if not valid:
        print("  无有效数据, 退出")
        sys.exit(1)

    # 打印将执行的操作 (前 10 条)
    print(f"\n将处理 (前 10 条):")
    for r in valid[:10]:
        print(f"  {r['scope']:8} {r['scope_value']:30} → {r['owner_name']}")
    if len(valid) > 10:
        print(f"  …还有 {len(valid) - 10} 条")

    if not args.yes:
        print(f"\n[DRY-RUN] 加 --yes 实际写入")
        return

    print(f"\n正在写入数据库…")
    inserted, updated = asyncio.run(do_upsert(valid))
    print(f"  ✓ 新增 {inserted} 条, 更新 {updated} 条, 错误 {len(errors)} 条")


if __name__ == "__main__":
    main()
